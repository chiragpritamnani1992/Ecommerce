import openpyxl


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readDataFile(file, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(file, sheetname, rownum, columnum)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnum).value


def writeData(file, sheetname, rownum, columnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnum).value = data
    workbook.save(file)
